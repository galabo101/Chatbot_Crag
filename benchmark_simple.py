"""
Benchmark Đơn Giản - Đánh giá PASS/FAIL
Tạo 2 file output:
1. benchmark_results_full.xlsx: question + answer (đánh giá thủ công)
2. benchmark_results_eval.xlsx: question + PASS/FAIL + thống kê (LLM đánh giá)
"""

import time
import sys
from pathlib import Path
from datetime import datetime
from groq import Groq
import os
from dotenv import load_dotenv

load_dotenv()

# Setup path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from src.pipeline import RAGPipeline

# Check for openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl not installed. Will use CSV with BOM instead.")


class SimpleBenchmark:
    def __init__(self, questions_file: str = "benchmark_questions.txt"):
        self.questions_file = questions_file
        self.questions = self.load_questions()
        self.groq_client = Groq(api_key=os.getenv("GROQ_API_KEY"))
        
        print(f"📋 Loaded {len(self.questions)} questions")
        
        # Load pipeline
        print("🔧 Loading RAG pipeline...")
        from sentence_transformers import SentenceTransformer
        embedding_model = SentenceTransformer("google/embeddinggemma-300m")
        self.pipeline = RAGPipeline(
            model_type="gemma",
            verbose=False,
            preloaded_model=embedding_model
        )
        print("✅ Pipeline ready")
    
    def load_questions(self) -> list:
        """Load câu hỏi từ file txt"""
        questions = []
        with open(self.questions_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    questions.append(line)
        return questions
    
    def evaluate_answer(self, question: str, answer: str) -> tuple:
        """
        Dùng LLM đánh giá: PASS, FAIL_WRONG hoặc FAIL_NO_DATA
        Returns: (result, fail_reason)
            - result: "PASS" hoặc "FAIL"
            - fail_reason: None, "WRONG", hoặc "NO_DATA"
        """
        prompt = f"""Đánh giá câu trả lời của chatbot tư vấn tuyển sinh Đại học Bình Dương.

CÂU HỎI: {question}
CÂU TRẢ LỜI: {answer[:1500]}

Tiêu chí đánh giá:
- PASS: Câu trả lời có thông tin liên quan, hữu ích, hoặc trả lời đúng câu hỏi
- FAIL_WRONG: Câu trả lời SAI, không chính xác, hoặc đưa thông tin không đúng
- FAIL_NO_DATA: Chatbot nói "không tìm thấy", "không có thông tin", "không biết", hoặc câu trả lời quá chung chung không cụ thể

CHỈ trả về một trong ba: PASS, FAIL_WRONG, FAIL_NO_DATA"""

        try:
            response = self.groq_client.chat.completions.create(
                messages=[{"role": "user", "content": prompt}],
                model="llama-3.1-8b-instant",
                temperature=0.1,
                max_tokens=20
            )
            result = response.choices[0].message.content.strip().upper()
            
            if "PASS" in result:
                return ("PASS", None)
            elif "NO_DATA" in result or "NODATA" in result:
                return ("FAIL", "NO_DATA")
            elif "WRONG" in result:
                return ("FAIL", "WRONG")
            else:
                # Default to WRONG if just FAIL
                return ("FAIL", "WRONG")
        except Exception as e:
            print(f"   ⚠️ Eval error: {e}")
            return ("FAIL", "ERROR")
    
    def detect_action(self, response: dict) -> str:
        """Detect action taken from pipeline response"""
        graded_stats = response.get("graded_stats", {})
        
        # Check based on graded stats
        correct = graded_stats.get("correct", 0)
        ambiguous = graded_stats.get("ambiguous", 0)
        incorrect = graded_stats.get("incorrect", 0)
        
        if correct >= 1:
            return "CORRECT"
        elif ambiguous >= 1 and correct == 0:
            return "KNOWLEDGE_REFINEMENT"
        elif incorrect >= 1 and correct == 0 and ambiguous == 0:
            return "WEB_SEARCH"
        else:
            return "UNKNOWN"
    
    def save_excel(self, results_full: list, results_eval: list, stats: dict, timestamp: str):
        """Lưu kết quả dạng Excel"""
        
        # ===== FILE 1: Full Results (question + answer) =====
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Kết quả chi tiết"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Headers
        ws1["A1"] = "STT"
        ws1["B1"] = "Câu hỏi"
        ws1["C1"] = "Câu trả lời"
        
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data
        for i, row in enumerate(results_full, start=2):
            ws1[f"A{i}"] = i - 1
            ws1[f"B{i}"] = row["question"]
            ws1[f"C{i}"] = row["answer"]
            ws1[f"B{i}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws1[f"C{i}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        # Column widths
        ws1.column_dimensions["A"].width = 5
        ws1.column_dimensions["B"].width = 50
        ws1.column_dimensions["C"].width = 80
        
        full_file = f"benchmark_results_full_{timestamp}.xlsx"
        wb1.save(full_file)
        
        # ===== FILE 2: Eval Results (question + result + stats) =====
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Đánh giá"
        
        ws2["A1"] = "STT"
        ws2["B1"] = "Câu hỏi"
        ws2["C1"] = "Kết quả"
        ws2["D1"] = "Action"
        ws2["E1"] = "Thời gian"
        
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for i, row in enumerate(results_eval, start=2):
            ws2[f"A{i}"] = i - 1
            ws2[f"B{i}"] = row["question"]
            ws2[f"C{i}"] = row["result"]
            ws2[f"D{i}"] = row["action"]
            ws2[f"E{i}"] = f"{row['response_time']:.2f}s"
            
            # Color based on result
            if row["result"] == "PASS":
                ws2[f"C{i}"].fill = pass_fill
            else:
                ws2[f"C{i}"].fill = fail_fill
        
        # Add summary section
        summary_row = len(results_eval) + 3
        ws2[f"A{summary_row}"] = "THỐNG KÊ"
        ws2[f"A{summary_row}"].font = Font(bold=True, size=12)
        
        ws2[f"A{summary_row + 1}"] = "Tổng câu hỏi:"
        ws2[f"B{summary_row + 1}"] = stats["total"]
        
        ws2[f"A{summary_row + 2}"] = "PASS:"
        ws2[f"B{summary_row + 2}"] = stats["passed"]
        ws2[f"B{summary_row + 2}"].fill = pass_fill
        
        ws2[f"A{summary_row + 3}"] = "FAIL:"
        ws2[f"B{summary_row + 3}"] = stats["failed"]
        ws2[f"B{summary_row + 3}"].fill = fail_fill
        
        ws2[f"A{summary_row + 4}"] = "  - Trả lời sai:"
        ws2[f"B{summary_row + 4}"] = stats["fail_wrong"]
        
        ws2[f"A{summary_row + 5}"] = "  - Thiếu dữ liệu:"
        ws2[f"B{summary_row + 5}"] = stats["fail_no_data"]
        
        ws2[f"A{summary_row + 6}"] = "  - Lỗi hệ thống:"
        ws2[f"B{summary_row + 6}"] = stats["fail_error"]
        
        ws2[f"A{summary_row + 7}"] = "Accuracy:"
        ws2[f"B{summary_row + 7}"] = f"{stats['accuracy']:.1f}%"
        ws2[f"B{summary_row + 7}"].font = Font(bold=True)
        
        # Action stats
        ws2[f"A{summary_row + 9}"] = "THỐNG KÊ ACTION"
        ws2[f"A{summary_row + 9}"].font = Font(bold=True, size=12)
        
        ws2[f"A{summary_row + 10}"] = "CORRECT:"
        ws2[f"B{summary_row + 10}"] = stats["action_correct"]
        
        ws2[f"A{summary_row + 11}"] = "KNOWLEDGE_REFINEMENT:"
        ws2[f"B{summary_row + 11}"] = stats["action_refinement"]
        
        ws2[f"A{summary_row + 12}"] = "WEB_SEARCH:"
        ws2[f"B{summary_row + 12}"] = stats["action_websearch"]
        
        ws2[f"A{summary_row + 13}"] = "UNKNOWN:"
        ws2[f"B{summary_row + 13}"] = stats["action_unknown"]
        
        ws2[f"A{summary_row + 15}"] = "THỐNG KÊ THỜI GIAN"
        ws2[f"A{summary_row + 15}"].font = Font(bold=True, size=12)
        
        ws2[f"A{summary_row + 16}"] = "Thời gian TB:"
        ws2[f"B{summary_row + 16}"] = f"{stats['avg_response_time']:.2f}s"
        ws2[f"B{summary_row + 16}"].font = Font(bold=True)

        
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 60
        ws2.column_dimensions["C"].width = 10
        ws2.column_dimensions["D"].width = 25
        ws2.column_dimensions["E"].width = 12
        
        eval_file = f"benchmark_results_eval_{timestamp}.xlsx"
        wb2.save(eval_file)
        
        return full_file, eval_file
    
    def save_csv_with_bom(self, results_full: list, results_eval: list, stats: dict, timestamp: str):
        """Fallback: Lưu CSV với BOM để Excel đọc được UTF-8"""
        import csv
        
        # File 1: Full results
        full_file = f"benchmark_results_full_{timestamp}.csv"
        with open(full_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["STT", "Câu hỏi", "Câu trả lời"])
            for i, row in enumerate(results_full, 1):
                writer.writerow([i, row["question"], row["answer"]])
        
        # File 2: Eval results
        eval_file = f"benchmark_results_eval_{timestamp}.csv"
        with open(eval_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["STT", "Câu hỏi", "Kết quả", "Action", "Thời gian"])
            for i, row in enumerate(results_eval, 1):
                writer.writerow([i, row["question"], row["result"], row["action"], f"{row['response_time']:.2f}s"])
            
            # Empty rows then stats
            writer.writerow([])
            writer.writerow(["THỐNG KÊ"])
            writer.writerow(["Tổng câu hỏi", stats["total"]])
            writer.writerow(["PASS", stats["passed"]])
            writer.writerow(["FAIL", stats["failed"]])
            writer.writerow(["Accuracy", f"{stats['accuracy']:.1f}%"])
            writer.writerow([])
            writer.writerow(["CORRECT", stats["action_correct"]])
            writer.writerow(["KNOWLEDGE_REFINEMENT", stats["action_refinement"]])
            writer.writerow(["WEB_SEARCH", stats["action_websearch"]])
            writer.writerow(["UNKNOWN", stats["action_unknown"]])
            writer.writerow([])
            writer.writerow(["THỐNG KÊ THỜI GIAN"])
            writer.writerow(["Thời gian TB", f"{stats['avg_response_time']:.2f}s"])
        
        return full_file, eval_file
    
    def run(self, limit: int = None, delay: int = 30):
        """
        Chạy benchmark
        Args:
            limit: Số câu hỏi tối đa (None = tất cả)
            delay: Delay giữa các câu hỏi (giây)
        """
        questions_to_run = self.questions[:limit] if limit else self.questions
        total = len(questions_to_run)
        
        print(f"\n{'='*60}")
        print(f"BENCHMARK: {total} câu hỏi")
        print(f"{'='*60}\n")
        
        results_full = []
        results_eval = []
        response_times = []
        
        passed = 0
        failed = 0
        fail_wrong = 0
        fail_no_data = 0
        fail_error = 0
        action_correct = 0
        action_refinement = 0
        action_websearch = 0
        action_unknown = 0
        
        for i, question in enumerate(questions_to_run, 1):
            print(f"[{i}/{total}] {question[:60]}...")
            
            try:
                # Chạy pipeline
                start_time = time.time()
                response = self.pipeline.run(question, user_id="benchmark_user")
                response_time = time.time() - start_time
                response_times.append(response_time)
                
                answer = response.get("answer", "")
                
                # Detect action
                action = self.detect_action(response)
                if action == "CORRECT":
                    action_correct += 1
                elif action == "KNOWLEDGE_REFINEMENT":
                    action_refinement += 1
                elif action == "WEB_SEARCH":
                    action_websearch += 1
                else:
                    action_unknown += 1
                
                # LLM đánh giá
                result, fail_reason = self.evaluate_answer(question, answer)
                
                if result == "PASS":
                    passed += 1
                    print(f"   ✅ PASS | Action: {action} | Time: {response_time:.2f}s")
                else:
                    failed += 1
                    if fail_reason == "WRONG":
                        fail_wrong += 1
                        print(f"   ❌ FAIL (Sai) | Action: {action} | Time: {response_time:.2f}s")
                    elif fail_reason == "NO_DATA":
                        fail_no_data += 1
                        print(f"   ❌ FAIL (Thiếu dữ liệu) | Action: {action} | Time: {response_time:.2f}s")
                    else:
                        fail_error += 1
                        print(f"   ❌ FAIL (Error) | Action: {action} | Time: {response_time:.2f}s")
                
                # Tạo result string cho output
                result_str = result if result == "PASS" else f"FAIL_{fail_reason}"
                
                results_full.append({
                    "question": question,
                    "answer": answer
                })
                results_eval.append({
                    "question": question,
                    "result": result_str,
                    "fail_reason": fail_reason,
                    "action": action,
                    "response_time": response_time
                })
                
            except Exception as e:
                print(f"   ❌ ERROR: {e}")
                failed += 1
                fail_error += 1
                action_unknown += 1
                results_full.append({
                    "question": question,
                    "answer": f"ERROR: {e}"
                })
                results_eval.append({
                    "question": question,
                    "result": "FAIL_ERROR",
                    "fail_reason": "ERROR",
                    "action": "ERROR",
                    "response_time": 0
                })
            
            # Delay để tránh rate limit
            if i < total:
                time.sleep(delay)
        
        # Tính accuracy và avg response time
        accuracy = (passed / total) * 100 if total > 0 else 0
        avg_response_time = sum(response_times) / len(response_times) if response_times else 0
        
        stats = {
            "total": total,
            "passed": passed,
            "failed": failed,
            "fail_wrong": fail_wrong,
            "fail_no_data": fail_no_data,
            "fail_error": fail_error,
            "accuracy": accuracy,
            "action_correct": action_correct,
            "action_refinement": action_refinement,
            "action_websearch": action_websearch,
            "action_unknown": action_unknown,
            "avg_response_time": avg_response_time
        }
        
        # Lưu kết quả
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if EXCEL_AVAILABLE:
            full_file, eval_file = self.save_excel(results_full, results_eval, stats, timestamp)
        else:
            full_file, eval_file = self.save_csv_with_bom(results_full, results_eval, stats, timestamp)
        
        # In kết quả
        print(f"\n{'='*60}")
        print("KẾT QUẢ BENCHMARK")
        print(f"{'='*60}")
        print(f"📊 Tổng câu hỏi: {total}")
        print(f"✅ PASS: {passed}")
        print(f"❌ FAIL: {failed}")
        print(f"   - Trả lời sai: {fail_wrong}")
        print(f"   - Thiếu dữ liệu: {fail_no_data}")
        print(f"   - Lỗi hệ thống: {fail_error}")
        print(f"🎯 Accuracy: {accuracy:.1f}%")
        print(f"\n📈 Thống kê Action:")
        print(f"   - CORRECT: {action_correct}")
        print(f"   - KNOWLEDGE_REFINEMENT: {action_refinement}")
        print(f"   - WEB_SEARCH: {action_websearch}")
        print(f"   - UNKNOWN: {action_unknown}")
        print(f"\n⏱️ Thời gian phản hồi TB: {avg_response_time:.2f}s")
        print(f"\n📁 File kết quả:")
        print(f"   - {full_file} (đánh giá thủ công)")
        print(f"   - {eval_file} (đánh giá tự động)")
        
        return stats


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Simple Benchmark PASS/FAIL")
    parser.add_argument("--limit", type=int, default=None, help="Số câu hỏi tối đa")
    parser.add_argument("--delay", type=int, default=30, help="Delay giữa các câu (giây)")
    args = parser.parse_args()
    
    benchmark = SimpleBenchmark()
    benchmark.run(limit=args.limit, delay=args.delay)
